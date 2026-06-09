#!/usr/bin/env python3
"""Generate a formatted Excel workbook from gps_data_dict.py.

The workbook creates one worksheet per region (R01, R02, etc.) and includes:
- a title
- a table of pole rows
- a summary line with total poles and total lights
"""
from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from gps_data_dict import gps_data


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
TITLE_FILL = PatternFill("solid", fgColor="0F243E")
SUMMARY_FILL = PatternFill("solid", fgColor="D9EAF7")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
thin_side = Side(style="thin", color="B7C9D6")
THIN_BORDER = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)


def region_key(label: str) -> str:
    match = re.search(r"-R(\d+)-", label or "")
    if match:
        return f"R{match.group(1)}"
    return "OTHER"


def to_int_lights(value: str) -> int:
    if not value:
        return 0
    parts = re.findall(r"\d+", str(value))
    return sum(int(part) for part in parts)


def build_regions() -> dict[str, list[tuple[int, dict]]]:
    regions: dict[str, list[tuple[int, dict]]] = defaultdict(list)
    for pole_id, data in gps_data.items():
        regions[region_key(data.get("label", ""))].append((pole_id, data))
    return dict(sorted(regions.items(), key=lambda item: item[0]))


def style_table_sheet(ws, region: str, poles: list[tuple[int, dict]]) -> None:
    headers = ["Pole ID", "Label", "Lights", "Working", "Location", "Latitude", "Longitude"]

    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value = f"Pole Labels - {region}"
    title.fill = TITLE_FILL
    title.font = Font(color="FFFFFF", bold=True, size=14)
    title.alignment = Alignment(horizontal="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    row = 4
    total_lights = 0

    for pole_id, data in poles:
        values = [
            pole_id,
            data.get("label", ""),
            data.get("no of lights", ""),
            data.get("working", ""),
            data.get("location", "N/A"),
            data.get("latitude", ""),
            data.get("longitude", ""),
        ]

        total_lights += to_int_lights(values[2])

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx in (1, 3):
                cell.alignment = Alignment(horizontal="center")
        row += 1

    ws.cell(row=row + 1, column=1, value="Region Summary")
    ws.cell(row=row + 1, column=1).font = BOLD_FONT
    ws.cell(row=row + 1, column=1).fill = SUMMARY_FILL
    ws.cell(row=row + 1, column=2, value=f"Total poles: {len(poles)}")
    ws.cell(row=row + 1, column=2).fill = SUMMARY_FILL
    ws.cell(row=row + 1, column=3, value=f"Total lights: {total_lights}")
    ws.cell(row=row + 1, column=3).fill = SUMMARY_FILL

    for col_idx, width in {
        1: 10,
        2: 28,
        3: 10,
        4: 10,
        5: 34,
        6: 14,
        7: 14,
    }.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A4"


def build_summary_sheet(wb: Workbook, regions: dict[str, list[tuple[int, dict]]]) -> None:
    ws = wb.create_sheet("Summary", 0)
    ws["A1"] = "Pole Summary"
    ws["A1"].font = Font(bold=True, size=14)

    summary_headers = ["Region", "Pole Count", "Total Lights"]
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    row = 4
    grand_poles = 0
    grand_lights = 0
    for region, poles in regions.items():
        region_lights = sum(to_int_lights(data.get("no of lights", "0")) for _, data in poles)
        grand_poles += len(poles)
        grand_lights += region_lights
        ws.cell(row=row, column=1, value=region)
        ws.cell(row=row, column=2, value=len(poles))
        ws.cell(row=row, column=3, value=region_lights)
        row += 1

    ws.cell(row=row + 1, column=1, value="Grand Total")
    ws.cell(row=row + 1, column=2, value=grand_poles)
    ws.cell(row=row + 1, column=3, value=grand_lights)
    for col in range(1, 4):
        ws.cell(row=row + 1, column=col).fill = SUMMARY_FILL
        ws.cell(row=row + 1, column=col).font = BOLD_FONT

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14


def main() -> int:
    regions = build_regions()
    wb = Workbook()

    default_sheet = wb.active
    wb.remove(default_sheet)

    build_summary_sheet(wb, regions)

    for region, poles in regions.items():
        ws = wb.create_sheet(region)
        poles = sorted(poles, key=lambda item: item[0])
        style_table_sheet(ws, region, poles)

    out_path = Path("poles_by_region.xlsx")
    wb.save(out_path)
    print(f"Created {out_path}")
    print(f"Regions: {len(regions)}")
    print(f"Total poles: {len(gps_data)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
